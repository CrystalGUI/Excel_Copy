import streamlit as st
import openpyxl
from io import BytesIO
from copy import copy  # 用于复制格式

# 设置网页标题和图标
st.set_page_config(
    page_title="Excel权限解除工具",
    page_icon="📊",
    layout="centered"
)

# 添加说明和注意事项
st.title("🔓 Excel权限解除工具")
st.markdown("""
> ⚠️ **注意**：  
> - 仅支持 `.xlsx` 文件（不支持旧版 `.xls`）  
> - 会保留所有单元格值和基础格式（字体/颜色/边框）  
> - 宏/VBA等高级功能无法复制  
""")

# 文件上传区域
uploaded_file = st.file_uploader(
    "拖拽需要解除权限的Excel文件到这里",
    type=["xlsx"],
    accept_multiple_files=False,
    help="最大支持100MB的文件"
)


def copy_excel_with_format(source):
    """复制Excel并保留格式"""
    source_workbook = openpyxl.load_workbook(source, read_only=True)
    source_sheet = source_workbook.active

    new_workbook = openpyxl.Workbook()
    new_sheet = new_workbook.active

    # 复制单元格值和格式
    for row in source_sheet.iter_rows():
        for cell in row:
            new_cell = new_sheet[cell.coordinate]
            new_cell.value = cell.value

            # 复制样式（如果存在）
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.border = copy(cell.border)
                new_cell.fill = copy(cell.fill)
                new_cell.number_format = cell.number_format
                new_cell.alignment = copy(cell.alignment)

    return new_workbook


if uploaded_file:
    try:
        with st.spinner("正在处理文件，请稍候..."):
            # 处理文件
            new_workbook = copy_excel_with_format(uploaded_file)

            # 生成下载文件
            output = BytesIO()
            new_workbook.save(output)
            output.seek(0)

            # 显示成功信息
            st.success("✅ 文件处理完成！")

            # 添加下载按钮
            st.download_button(
                label="⬇️ 下载可编辑副本",
                data=output,
                file_name=f"unlocked_{uploaded_file.name}",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                help="点击下载后，请检查文件内容是否完整"
            )

    except Exception as e:
        st.error(f"❌ 处理失败: {str(e)}")
        st.info("请检查文件是否受密码保护或损坏")

# 添加页脚
st.markdown("---")
st.caption("ℹ️ 技术支持：Streamlit + OpenPyXL | 隐私声明：文件不会存储在服务器")